import openpyxl
import string
from io import BytesIO
from copy import copy

from django.db.models import Prefetch, Sum, FloatField
from django.db.models.functions import Coalesce

from src.payment_receipts.models import Receipt, ReceiptTemplate, ReceiptService
from src.system_settings.models import PaymentCredential
from src.cash_box.models import Transaction


def get_next_five_letters(letter):
    alphabet = list(string.ascii_uppercase)
    index = alphabet.index(letter)
    return alphabet[index + 1:index + 6]


class ReceiptExcelService:
    @classmethod
    def create_worksheet(cls, receipt_id, template_id):
        template = ReceiptTemplate.objects.get(pk=template_id)

        services_qs = (ReceiptService.objects
                       .select_related('service', 'unit')
                       .with_total_price())

        receipt = (Receipt.objects
                   .prefetch_related(Prefetch('services', queryset=services_qs))
                   .with_total_price()
                   .select_related('personal_account', 'tariff', 'flat', 'flat__owner', 'house', 'section')
                   .get(pk=receipt_id))

        paid = Transaction.objects.filter(
            receipt_id=receipt_id,
            type='expense',
            is_complete=True
        ).aggregate(paid_sum=Coalesce(Sum('amount'), 0.0, output_field=FloatField()))['paid_sum']

        credential = PaymentCredential.objects.first()

        data = {
            'number': receipt.no,
            'date': receipt.date.strftime("%d.%m.%Y"),
            'period_from': receipt.period_from.strftime("%d.%m.%Y"),
            'period_to': receipt.period_to.strftime("%d.%m.%Y"),
            'company_name': credential.name,
            'company_information': credential.information,
            'full_name': f"{receipt.flat.owner.last_name} "
                         f"{receipt.flat.owner.first_name} "
                         f"{receipt.flat.owner.middle_name}",
            'phone_number': receipt.flat.owner.phone_number,
            'email': receipt.flat.owner.email,
            'address': f"{receipt.house.address}, {receipt.section}, кв. № {receipt.flat.no}",
            'personal_account': receipt.personal_account.no,
            'sum': receipt.total_price,
            'paid': paid,
            'debt': receipt.total_price - paid
        }

        workbook = openpyxl.load_workbook(filename=str(template.file.file))
        sheet = workbook.active

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value in data.keys():
                    sheet[cell.coordinate] = data[cell.value]

        service_row = None
        counter_col = None

        font = None
        border = None
        fill = None
        alignment = None

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == 'counter':
                    service_row = cell.row
                    counter_col = cell.column_letter
                    font = copy(cell.font)
                    border = copy(cell.border)
                    fill = copy(cell.fill)
                    alignment = copy(cell.alignment)
                    break

        service_col, value_col, unit_col, price_col, total_col = get_next_five_letters(counter_col)

        for index, service in enumerate(receipt.services.all()):
            if index > 0:
                sheet.insert_rows(service_row)

            sheet[f'{counter_col}{service_row}'] = index + 1
            sheet[f'{service_col}{service_row}'] = service.service.name
            sheet[f'{value_col}{service_row}'] = service.value
            sheet[f'{unit_col}{service_row}'] = service.unit.unit
            sheet[f'{price_col}{service_row}'] = service.price
            sheet[f'{total_col}{service_row}'] = service.total_price

            sheet[f'{counter_col}{service_row}'].font = font
            sheet[f'{counter_col}{service_row}'].border = border
            sheet[f'{counter_col}{service_row}'].fill = fill
            sheet[f'{counter_col}{service_row}'].alignment = alignment

            sheet[f'{service_col}{service_row}'].font = font
            sheet[f'{service_col}{service_row}'].border = border
            sheet[f'{service_col}{service_row}'].fill = fill
            sheet[f'{service_col}{service_row}'].alignment = alignment

            sheet[f'{value_col}{service_row}'].font = font
            sheet[f'{value_col}{service_row}'].border = border
            sheet[f'{value_col}{service_row}'].fill = fill
            sheet[f'{value_col}{service_row}'].alignment = alignment

            sheet[f'{unit_col}{service_row}'].font = font
            sheet[f'{unit_col}{service_row}'].border = border
            sheet[f'{unit_col}{service_row}'].fill = fill
            sheet[f'{unit_col}{service_row}'].alignment = alignment

            sheet[f'{price_col}{service_row}'].font = font
            sheet[f'{price_col}{service_row}'].border = border
            sheet[f'{price_col}{service_row}'].fill = fill
            sheet[f'{price_col}{service_row}'].alignment = alignment

            sheet[f'{total_col}{service_row}'].font = font
            sheet[f'{total_col}{service_row}'].border = border
            sheet[f'{total_col}{service_row}'].fill = fill
            sheet[f'{total_col}{service_row}'].alignment = alignment

            service_row += 1

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output, f"Receipt_{receipt.no}.xlsx"
