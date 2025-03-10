import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO


class TransactionExcelService:
    @classmethod
    def create_worksheet(cls, transaction):
        wb = openpyxl.Workbook()
        ws = wb.active

        cls.__set_column_width(ws)
        cls.__set_borders(ws)
        cls.__set_alignment(ws)
        cls.__append_header(ws, transaction)
        cls.__append_data(ws, transaction)
        cls.__set_font(ws, transaction)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def __set_column_width(ws):
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 40

    @staticmethod
    def __set_borders(ws):
        border = Border(
            left=Side(border_style='thin'),
            right=Side(border_style='thin'),
            top=Side(border_style='thin'),
            bottom=Side(border_style='thin')
        )

        for row in range(1, 9):
            for col in ('A', 'B'):
                ws[f"{col}{row}"].border = border

    @staticmethod
    def __set_alignment(ws):
        head_alignment = Alignment(horizontal="center", vertical="center")
        data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)

        ws["A1"].alignment = head_alignment
        ws["B1"].alignment = head_alignment

        for row in range(2, 9):
            for col in ('A', 'B'):
                ws[f"{col}{row}"].alignment = data_alignment

    @staticmethod
    def __append_data(ws, transaction):
        ws["A2"].value = "Власник квартири"
        ws["B2"].value = str(transaction.owner) if transaction.owner else "Не вказано"
        ws["A3"].value = "Особовий рахунок"
        ws["B3"].value = transaction.personal_account.no if transaction.personal_account else "Не вказано"
        ws["A4"].value = "Стаття"
        ws["B4"].value = transaction.payment_item.name if transaction.payment_item else "Не вказано"
        ws["A5"].value = "Квитанція"
        ws["B5"].value = transaction.receipt.no if transaction.receipt else "Не вказано"
        ws["A6"].value = "Менеджер"
        ws["B6"].value = str(transaction.manager) if transaction.manager else "Не вказано"
        ws["A7"].value = "Сума"
        ws["B7"].value = transaction.amount
        ws["A8"].value = "Коментар"
        ws["B8"].value = transaction.comment

    @staticmethod
    def __append_header(ws, transaction):
        title = str(transaction).title()
        ws.merge_cells("A1:B1")
        ws["A1"] = title

    @staticmethod
    def __set_font(ws, transaction):
        name = 'Calibri'
        color = '000000'
        size = 12
        bold = False
        # header
        ws[f"A1"].font = Font(size=14, bold=True, name=name, color=color)

        for row in range(2, 9):
            for col in ('A', 'B'):
                # if amount value
                if row == 7 and col == 'B':
                    amount_color = "00FF00" if transaction.type == "income" else "FF0000"
                    ws[f"{col}{row}"].font = Font(size=size, bold=bold, name=name, color=amount_color)
                else:
                    ws[f"{col}{row}"].font = Font(size=size, bold=bold, name=name, color=color)
