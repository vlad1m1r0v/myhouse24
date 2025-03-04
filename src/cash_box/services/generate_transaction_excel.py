import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO


def generate_transaction_excel(transaction):
    data = [
        ("Власник квартири", str(transaction.owner) if transaction.owner else "Не вказано"),
        ("Особовий рахунок", str(transaction.personal_account) if transaction.personal_account else "Не вказано"),
        ("Стаття", str(transaction.payment_item) if transaction.payment_item else "Не вказано"),
        ("Квитанція", str(transaction.receipt) if transaction.receipt else "Не вказано"),
        ("Менеджер", str(transaction.manager) if transaction.manager else "Не вказано"),
        ("Сума", transaction.amount),
        ("Коментар", transaction.comment)
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    # set width of columns
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40

    border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )

    for row in range(1, len(data) + 3):
        ws.row_dimensions[row].height = 24

        for col in range(1, 3):
            cell = ws.cell(row=row, column=col)

            if row < 3:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(size=14, bold=True)

            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.font = Font(size=12)

            cell.border = border

    title = str(transaction).title()
    ws.merge_cells("A1:B1")
    ws["A1"] = title

    ws['A2'].value = 'Поле'
    ws['B2'].value = 'Значення'

    for row_idx, (field, value) in enumerate(data, start=3):
        ws.cell(row=row_idx, column=1).value = field
        ws.cell(row=row_idx, column=2).value = value

        if field == "Сума":
            fill_color = "00FF00" if transaction.type == "income" else "FF0000"
            ws.cell(row=row_idx, column=2).font = Font(color=fill_color)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
