from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO


class CashBoxExcelService:
    @classmethod
    def create_worksheet(cls, transactions):
        wb = openpyxl.Workbook()
        ws = wb.active

        cls.__set_column_width(ws)
        cls.__append_header(ws)

        for row_idx, transaction in enumerate(transactions, start=2):
            cls.__set_alignment(ws, row_idx)
            cls.__set_font(ws, row_idx)
            cls.__set_borders(ws, row_idx)
            cls.__append_record(ws, row_idx, transaction)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def __set_column_width(ws):
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 25
        ws.column_dimensions["F"].width = 25
        ws.column_dimensions["G"].width = 20
        ws.column_dimensions["H"].width = 25
        ws.column_dimensions["I"].width = 25

    @staticmethod
    def __set_alignment(ws, row_idx):
        alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)

        ws[f"A{row_idx}"].alignment = alignment
        ws[f"B{row_idx}"].alignment = alignment
        ws[f"C{row_idx}"].alignment = alignment
        ws[f"D{row_idx}"].alignment = alignment
        ws[f"E{row_idx}"].alignment = alignment
        ws[f"F{row_idx}"].alignment = alignment
        ws[f"G{row_idx}"].alignment = alignment
        ws[f"H{row_idx}"].alignment = alignment
        ws[f"I{row_idx}"].alignment = alignment

    @staticmethod
    def __set_borders(ws, row_idx):
        border = Border(
            left=Side(border_style='thin'),
            right=Side(border_style='thin'),
            top=Side(border_style='thin'),
            bottom=Side(border_style='thin')
        )

        ws[f"A{row_idx}"].border = border
        ws[f"B{row_idx}"].border = border
        ws[f"C{row_idx}"].border = border
        ws[f"D{row_idx}"].border = border
        ws[f"E{row_idx}"].border = border
        ws[f"F{row_idx}"].border = border
        ws[f"G{row_idx}"].border = border
        ws[f"H{row_idx}"].border = border
        ws[f"I{row_idx}"].border = border

    @staticmethod
    def __set_font(ws, row_idx):
        name = 'Calibri'
        size = 14 if row_idx == 1 else 12
        bold = row_idx == 1

        font = Font(name=name, size=size, bold=bold)

        ws[f"A{row_idx}"].font = font
        ws[f"B{row_idx}"].font = font
        ws[f"C{row_idx}"].font = font
        ws[f"D{row_idx}"].font = font
        ws[f"E{row_idx}"].font = font
        ws[f"F{row_idx}"].font = font
        ws[f"G{row_idx}"].font = font
        ws[f"H{row_idx}"].font = font
        ws[f"I{row_idx}"].font = font

    @classmethod
    def __append_header(cls, ws):
        cls.__set_alignment(ws, 1)
        cls.__set_font(ws, 1)
        cls.__set_borders(ws, 1)

        ws[f"A1"] = '№'
        ws[f"B1"].value = 'Дата'
        ws[f"C1"].value = 'Надходження / Витрата'
        ws[f"D1"].value = 'Статус'
        ws[f"E1"].value = 'Тип платежу'
        ws[f"F1"].value = 'Квитанція'
        ws[f"G1"].value = 'Сума'
        ws[f"H1"].value = 'Власник'
        ws[f"I1"].value = 'Особовий рахунок'

    @staticmethod
    def __append_record(ws, row_idx, transaction):
        ws[f"A{row_idx}"].value = transaction.no

        transaction_date = datetime.strptime(str(transaction.date), "%Y-%m-%d")
        formatted_date = transaction_date.strftime("%d.%m.%Y")
        ws[f"B{row_idx}"].value = formatted_date

        ws[f"C{row_idx}"].value = transaction.get_type_display()

        ws[f"D{row_idx}"].value = 'Проведена' if transaction.is_complete else 'Не проведена'

        ws[f"E{row_idx}"].value = transaction.payment_item.name

        ws[f"F{row_idx}"].value = transaction.receipt.no if transaction.receipt else "Не вказано"

        ws[f"G{row_idx}"].value = transaction.amount

        ws[f"H{row_idx}"].value = str(transaction.owner) if transaction.owner else "Не вказано"

        ws[f"I{row_idx}"].value = transaction.personal_account.no if transaction.personal_account else "Не вказано"
