import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO


class AccountExcelService:
    @classmethod
    def create_worksheet(cls, accounts):
        wb = openpyxl.Workbook()
        ws = wb.active

        cls.__set_column_width(ws)
        cls.__append_header(ws)

        for row_idx, account in enumerate(accounts, start=2):
            cls.__set_alignment(ws, row_idx)
            cls.__set_font(ws, row_idx)
            cls.__set_borders(ws, row_idx)
            cls.__append_record(ws, row_idx, account)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def __set_column_width(ws):
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 25
        ws.column_dimensions["F"].width = 25
        ws.column_dimensions["G"].width = 25

    @staticmethod
    def __set_alignment(ws, row_idx):
        alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)

        ws[f"A{row_idx}"].alignment = alignment
        ws[f"B{row_idx}"].alignment = alignment
        ws[f"C{row_idx}"].alignment = alignment
        ws[f"D{row_idx}"].alignment = alignment
        ws[f"E{row_idx}"].alignment = alignment
        ws[f"F{row_idx}"].alignment = alignment
        ws[f"G{row_idx}"].alignment = alignment

    @staticmethod
    def __set_borders(ws, row_idx):
        border = Border(
            left=Side(border_style='thin'),
            right=Side(border_style='thin'),
            top=Side(border_style='thin'),
            bottom=Side(border_style='thin')
        )

        ws[f"A{row_idx}"].border = border
        ws[f"B{row_idx}"].border = border
        ws[f"C{row_idx}"].border = border
        ws[f"D{row_idx}"].border = border
        ws[f"E{row_idx}"].border = border
        ws[f"F{row_idx}"].border = border
        ws[f"G{row_idx}"].border = border

    @staticmethod
    def __set_font(ws, row_idx):
        name = 'Calibri'
        size = 14 if row_idx == 1 else 12
        bold = row_idx == 1

        font = Font(name=name, size=size, bold=bold)

        ws[f"A{row_idx}"].font = font
        ws[f"B{row_idx}"].font = font
        ws[f"C{row_idx}"].font = font
        ws[f"D{row_idx}"].font = font
        ws[f"E{row_idx}"].font = font
        ws[f"F{row_idx}"].font = font
        ws[f"G{row_idx}"].font = font

    @classmethod
    def __append_header(cls, ws):
        cls.__set_alignment(ws, 1)
        cls.__set_font(ws, 1)
        cls.__set_borders(ws, 1)

        ws[f"A1"] = 'Особовий рахунок'
        ws[f"B1"].value = 'Статус'
        ws[f"C1"].value = 'Будинок'
        ws[f"D1"].value = 'Секція'
        ws[f"E1"].value = 'Квартира'
        ws[f"F1"].value = 'Власник'
        ws[f"G1"].value = 'Баланс'

    @staticmethod
    def __append_record(ws, row_idx, account):
        ws[f"A{row_idx}"].value = account.no

        ws[f"B{row_idx}"].value = account.get_status_display()

        ws[f"C{row_idx}"].value = account.house.name if account.house else "Не вказано"

        ws[f"D{row_idx}"].value = account.section.name if account.section else "Не вказано"

        ws[f"E{row_idx}"].value = account.flat.name if account.flat else "Не вказано"

        ws[f"F{row_idx}"].value = (f"{account.owner.last_name}"
                                   f" {account.owner.first_name}"
                                   f" {account.owner.middle_name}") if account.owner else "Не вказано"

        ws[f"G{row_idx}"].value = account.balance if account.balance else "Не вказано"